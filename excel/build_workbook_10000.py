"""
Builds excel/Ecommerce_Sales_Analysis_10000.xlsx — same 5-sheet architecture
as the prototype workbook (build_workbook.py), scaled to 10,000 rows.
Formulas (not hardcoded values) throughout; Raw_Data is converted to a
proper Excel Table so it auto-expands with future data loads.
"""
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, Reference
import csv

t0 = time.time()
SRC = "../data/ecommerce_data_10000.csv"
OUT = "Ecommerce_Sales_Analysis_10000.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
LABEL_FONT = Font(name=FONT, bold=True, size=10)
NORMAL_FONT = Font(name=FONT, size=10)
KPI_FILL = PatternFill("solid", fgColor="EAF1F8")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------
# 1. RAW_DATA (fast bulk append, header styled, converted to Table)
# ---------------------------------------------------------------
ws = wb.create_sheet("Raw_Data")
with open(SRC) as f:
    reader = csv.reader(f)
    header = next(reader)
    ws.append(header)
    numeric_int_cols = {"Quantity","Age","Pages_Viewed","Order_Completed","Shipping_Days"}
    numeric_float_cols = {"Unit_Price","Discount","Cost","Session_Duration"}
    idx_int = [i for i,h in enumerate(header) if h in numeric_int_cols]
    idx_float = [i for i,h in enumerate(header) if h in numeric_float_cols]
    n = 0
    for row in reader:
        for i in idx_int:
            row[i] = int(row[i])
        for i in idx_float:
            row[i] = float(row[i])
        ws.append(row)
        n += 1

N = n
LAST_ROW = N + 1
for c, h in enumerate(header, 1):
    cell = ws.cell(row=1, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(c)].width = max(12, len(h) + 2)
ws.freeze_panes = "A2"

table_ref = f"A1:{get_column_letter(len(header))}{LAST_ROW}"
tbl = Table(displayName="RawDataTable", ref=table_ref)
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

col = {name: get_column_letter(i + 1) for i, name in enumerate(header)}
print(f"Raw_Data written: {N} rows in {time.time()-t0:.1f}s")

# ---------------------------------------------------------------
# 2. DATA_QUALITY
# ---------------------------------------------------------------
dq = wb.create_sheet("Data_Quality")
dq["A1"] = "DATA QUALITY CHECKS — 10,000-ROW DATASET"
dq["A1"].font = TITLE_FONT
dq.merge_cells("A1:D1")

dq["A3"] = "Check"; dq["B3"] = "Result"; dq["C3"] = "Formula Used"
for cell in ("A3","B3","C3"):
    dq[cell].font = HEADER_FONT; dq[cell].fill = HEADER_FILL

checks = [
    ("Total rows", f"=COUNTA(Raw_Data!A2:A{LAST_ROW})", "COUNTA"),
    ("Duplicate Order_IDs", f'=SUMPRODUCT((COUNTIF(Raw_Data!A2:A{LAST_ROW},Raw_Data!A2:A{LAST_ROW})>1)*1)', "SUMPRODUCT + COUNTIF"),
    ("Missing Order_ID cells", f"=COUNTBLANK(Raw_Data!A2:A{LAST_ROW})", "COUNTBLANK"),
    ("Missing Customer_ID cells", f"=COUNTBLANK(Raw_Data!D2:D{LAST_ROW})", "COUNTBLANK"),
    ("Negative or zero Unit_Price", f"=COUNTIF(Raw_Data!L2:L{LAST_ROW},\"<=0\")", "COUNTIF"),
    ("Discount out of range (>0.5)", f"=COUNTIF(Raw_Data!M2:M{LAST_ROW},\">0.5\")", "COUNTIF"),
    ("Unique Customers", f"=SUMPRODUCT(1/COUNTIF(Raw_Data!D2:D{LAST_ROW},Raw_Data!D2:D{LAST_ROW}))", "SUMPRODUCT/COUNTIF distinct-count pattern"),
    ("Unique Products", f"=SUMPRODUCT(1/COUNTIF(Raw_Data!I2:I{LAST_ROW},Raw_Data!I2:I{LAST_ROW}))", "SUMPRODUCT/COUNTIF distinct-count pattern"),
]
r = 4
for label, formula, note in checks:
    dq.cell(row=r, column=1, value=label).font = NORMAL_FONT
    dq.cell(row=r, column=2, value=formula).font = NORMAL_FONT
    dq.cell(row=r, column=3, value=note).font = Font(name=FONT, size=9, italic=True, color="808080")
    r += 1

dq["A13"] = "Conclusion:"
dq["A13"].font = LABEL_FONT
dq["A14"] = ("10,000-row synthetic dataset generated with 0 duplicate Order_IDs, 0 missing "
             "critical fields, and referential integrity to 893 unique customers and 41 unique "
             "products (verified separately in sql/build_10000_db.py's orphan-key checks).")
dq["A14"].font = NORMAL_FONT
dq.merge_cells("A14:F14")
for col_letter, width in [("A",38),("B",16),("C",55)]:
    dq.column_dimensions[col_letter].width = width
print(f"Data_Quality written in {time.time()-t0:.1f}s")

# ---------------------------------------------------------------
# 3. CALCULATED_FIELDS (sample of first 500 rows with live formulas
#    referencing Raw_Data — full 10,000-row pattern documented; a sample
#    keeps the workbook responsive while proving every formula works)
# ---------------------------------------------------------------
SAMPLE_N = 500
cf = wb.create_sheet("Calculated_Fields")
cf["A1"] = f"Live formulas shown for first {SAMPLE_N:,} of {N:,} rows (same pattern extends to row {LAST_ROW})"
cf["A1"].font = Font(name=FONT, italic=True, size=9, color="808080")
cf.merge_cells("A1:K1")
headers = ["Order_ID","Order_Date","Order_Hour","Customer_ID","Customer_Segment",
           "Category (via lookup)","Revenue","Discount_Amount","Cost","Profit","Profit_Margin_%"]
for c, h in enumerate(headers, 1):
    cell = cf.cell(row=2, column=c, value=h)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for i in range(SAMPLE_N):
    r = i + 3
    src_r = i + 2
    cf.cell(row=r, column=1, value=f"=Raw_Data!{col['Order_ID']}{src_r}")
    cf.cell(row=r, column=2, value=f"=Raw_Data!{col['Order_Date']}{src_r}")
    cf.cell(row=r, column=3,
            value=f'=VALUE(LEFT(Raw_Data!{col["Order_Time"]}{src_r},FIND(":",Raw_Data!{col["Order_Time"]}{src_r})-1))')
    cf.cell(row=r, column=4, value=f"=Raw_Data!{col['Customer_ID']}{src_r}")
    cf.cell(row=r, column=5,
            value=f'=_xlfn.IFS(Raw_Data!{col["Customer_Type"]}{src_r}="New","New Customer",'
                  f'Raw_Data!{col["Customer_Type"]}{src_r}="Returning","Returning Customer",TRUE,"Unclassified")')
    cf.cell(row=r, column=6,
            value=f"=INDEX(Raw_Data!{col['Category']}:{col['Category']},MATCH(A{r},Raw_Data!{col['Order_ID']}:{col['Order_ID']},0))")
    cf.cell(row=r, column=7,
            value=f"=Raw_Data!{col['Quantity']}{src_r}*Raw_Data!{col['Unit_Price']}{src_r}*(1-Raw_Data!{col['Discount']}{src_r})")
    cf.cell(row=r, column=8,
            value=f"=Raw_Data!{col['Quantity']}{src_r}*Raw_Data!{col['Unit_Price']}{src_r}*Raw_Data!{col['Discount']}{src_r}")
    cf.cell(row=r, column=9, value=f"=Raw_Data!{col['Cost']}{src_r}")
    cf.cell(row=r, column=10, value=f"=G{r}-I{r}")
    cf.cell(row=r, column=11, value=f"=IFERROR(J{r}/G{r},0)")
    for c in range(7, 12):
        cf.cell(row=r, column=c).number_format = "#,##0.00" if c != 11 else "0.0%"
for c in range(1, 12):
    cf.column_dimensions[get_column_letter(c)].width = 16
cf.freeze_panes = "A3"
print(f"Calculated_Fields written ({SAMPLE_N} sample rows) in {time.time()-t0:.1f}s")

# ---------------------------------------------------------------
# 4. PIVOT_SUMMARY — SUMPRODUCT/COUNTIF cross-tabs over the FULL 10,000 rows
# ---------------------------------------------------------------
ps = wb.create_sheet("Pivot_Summary")
ps["A1"] = "PIVOT-STYLE SUMMARY (full 10,000-row dataset)"
ps["A1"].font = TITLE_FONT
ps.merge_cells("A1:E1")
ps["A2"] = "Tip: Select the RawDataTable and Insert > PivotTable in Excel for a native, drag-and-drop pivot."
ps["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

RD = "Raw_Data"
CAT_COL = col['Category']; REG_COL = col['Region']
CTYPE_COL = col['Customer_Type']; PAY_COL = col['Payment_Method']
QTY_COL = col['Quantity']; PRICE_COL = col['Unit_Price']; DISC_COL = col['Discount']
OC_COL = col['Order_Completed']

def revenue_sumifs(criteria_range, criteria_cell):
    return (f"=SUMPRODUCT(({RD}!{criteria_range}=${criteria_cell})*"
            f"({RD}!{OC_COL}2:{OC_COL}{LAST_ROW}=1)*"
            f"{RD}!{QTY_COL}2:{QTY_COL}{LAST_ROW}*{RD}!{PRICE_COL}2:{PRICE_COL}{LAST_ROW}*"
            f"(1-{RD}!{DISC_COL}2:{DISC_COL}{LAST_ROW}))")

ps["A4"] = "Revenue & Orders by Category"; ps["A4"].font = LABEL_FONT
ps["A5"] = "Category"; ps["B5"] = "Orders"; ps["C5"] = "Revenue"
for cell in ("A5","B5","C5"):
    ps[cell].font = HEADER_FONT; ps[cell].fill = HEADER_FILL

# distinct categories (small set, fine to read from python here for labels only)
import csv as _csv
with open(SRC) as f:
    _r = _csv.DictReader(f)
    _rows = list(_r)
categories = sorted(set(row["Category"] for row in _rows))
regions = sorted(set(row["Region"] for row in _rows))
methods = sorted(set(row["Payment_Method"] for row in _rows))

r = 6
for cat in categories:
    ps.cell(row=r, column=1, value=cat).font = NORMAL_FONT
    ps.cell(row=r, column=2,
            value=f'=COUNTIFS({RD}!{CAT_COL}2:{CAT_COL}{LAST_ROW},A{r},{RD}!{OC_COL}2:{OC_COL}{LAST_ROW},1)').font = NORMAL_FONT
    ps.cell(row=r, column=3, value=revenue_sumifs(f"{CAT_COL}2:{CAT_COL}{LAST_ROW}", f"A{r}")).font = NORMAL_FONT
    ps.cell(row=r, column=3).number_format = "#,##0.00"
    r += 1
cat_end = r - 1

start_b = cat_end + 3
ps.cell(row=start_b, column=1, value="Revenue & Orders by Region").font = LABEL_FONT
ps.cell(row=start_b+1, column=1, value="Region").font = HEADER_FONT
ps.cell(row=start_b+1, column=2, value="Orders").font = HEADER_FONT
ps.cell(row=start_b+1, column=3, value="Revenue").font = HEADER_FONT
for cc in range(1,4):
    ps.cell(row=start_b+1, column=cc).fill = HEADER_FILL
r = start_b + 2
for reg in regions:
    ps.cell(row=r, column=1, value=reg).font = NORMAL_FONT
    ps.cell(row=r, column=2,
            value=f'=COUNTIFS({RD}!{REG_COL}2:{REG_COL}{LAST_ROW},A{r},{RD}!{OC_COL}2:{OC_COL}{LAST_ROW},1)').font = NORMAL_FONT
    ps.cell(row=r, column=3, value=revenue_sumifs(f"{REG_COL}2:{REG_COL}{LAST_ROW}", f"A{r}")).font = NORMAL_FONT
    ps.cell(row=r, column=3).number_format = "#,##0.00"
    r += 1
reg_end = r - 1

start_c = reg_end + 3
ps.cell(row=start_c, column=1, value="New vs Returning Customers (per-order status)").font = LABEL_FONT
ps.cell(row=start_c+1, column=1, value="Customer_Type").font = HEADER_FONT
ps.cell(row=start_c+1, column=2, value="Orders").font = HEADER_FONT
ps.cell(row=start_c+1, column=3, value="Revenue").font = HEADER_FONT
for cc in range(1,4):
    ps.cell(row=start_c+1, column=cc).fill = HEADER_FILL
r = start_c + 2
for ctype in ["New","Returning"]:
    ps.cell(row=r, column=1, value=ctype).font = NORMAL_FONT
    ps.cell(row=r, column=2,
            value=f'=COUNTIFS({RD}!{CTYPE_COL}2:{CTYPE_COL}{LAST_ROW},A{r},{RD}!{OC_COL}2:{OC_COL}{LAST_ROW},1)').font = NORMAL_FONT
    ps.cell(row=r, column=3, value=revenue_sumifs(f"{CTYPE_COL}2:{CTYPE_COL}{LAST_ROW}", f"A{r}")).font = NORMAL_FONT
    ps.cell(row=r, column=3).number_format = "#,##0.00"
    r += 1
ctype_end = r - 1

start_d = ctype_end + 3
ps.cell(row=start_d, column=1, value="Orders by Payment Method").font = LABEL_FONT
ps.cell(row=start_d+1, column=1, value="Payment_Method").font = HEADER_FONT
ps.cell(row=start_d+1, column=2, value="Orders").font = HEADER_FONT
for cc in range(1,3):
    ps.cell(row=start_d+1, column=cc).fill = HEADER_FILL
r = start_d + 2
for m in methods:
    ps.cell(row=r, column=1, value=m).font = NORMAL_FONT
    ps.cell(row=r, column=2, value=f'=COUNTIFS({RD}!{PAY_COL}2:{PAY_COL}{LAST_ROW},A{r})').font = NORMAL_FONT
    r += 1
pay_end = r - 1

for c, w in zip("ABCDE", [24,12,16,16,16]):
    ps.column_dimensions[c].width = w

bar = BarChart()
bar.title = "Revenue by Category"
bar.y_axis.title = "Revenue (INR)"
bar.x_axis.title = "Category"
data_ref = Reference(ps, min_col=3, min_row=5, max_row=cat_end)
cats_ref = Reference(ps, min_col=1, min_row=6, max_row=cat_end)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.width = 16; bar.height = 9
ps.add_chart(bar, "E4")

pie = PieChart()
pie.title = "Revenue Share by Region"
data_ref2 = Reference(ps, min_col=3, min_row=start_b+1, max_row=reg_end)
cats_ref2 = Reference(ps, min_col=1, min_row=start_b+2, max_row=reg_end)
pie.add_data(data_ref2, titles_from_data=True)
pie.set_categories(cats_ref2)
pie.width = 16; pie.height = 9
ps.add_chart(pie, "E22")
print(f"Pivot_Summary written in {time.time()-t0:.1f}s")

# ---------------------------------------------------------------
# 5. KPI_DASHBOARD
# ---------------------------------------------------------------
kd = wb.create_sheet("KPI_Dashboard", 0)
kd["A1"] = "E-COMMERCE SALES PERFORMANCE — KPI DASHBOARD"
kd["A1"].font = Font(name=FONT, bold=True, size=16, color="1F4E78")
kd.merge_cells("A1:D1")
kd["A2"] = f"Scaled dataset: {N:,} orders | Jan–Dec 2026"
kd["A2"].font = Font(name=FONT, italic=True, size=10, color="808080")

OC = col['Order_Completed']; QC = col['Quantity']; PC = col['Unit_Price']; DC = col['Discount']; CC = col['Cost']
RET = col['Return_Status']; CUST = col['Customer_ID']

kpis = [
    ("Total Revenue",
     f"=SUMPRODUCT(({RD}!{OC}2:{OC}{LAST_ROW}=1)*{RD}!{QC}2:{QC}{LAST_ROW}*{RD}!{PC}2:{PC}{LAST_ROW}*(1-{RD}!{DC}2:{DC}{LAST_ROW}))",
     "#,##0.00"),
    ("Total Profit",
     f"=SUMPRODUCT(({RD}!{OC}2:{OC}{LAST_ROW}=1)*({RD}!{QC}2:{QC}{LAST_ROW}*{RD}!{PC}2:{PC}{LAST_ROW}*(1-{RD}!{DC}2:{DC}{LAST_ROW})-{RD}!{CC}2:{CC}{LAST_ROW}))",
     "#,##0.00"),
    ("Total Orders", f"=COUNTA({RD}!A2:A{LAST_ROW})", "#,##0"),
    ("Total Customers", f"=SUMPRODUCT(1/COUNTIF({RD}!{CUST}2:{CUST}{LAST_ROW},{RD}!{CUST}2:{CUST}{LAST_ROW}))", "#,##0"),
    ("Average Order Value", f"=B4/SUM({RD}!{OC}2:{OC}{LAST_ROW})", "#,##0.00"),
    ("Profit Margin %", "=B6/B4", "0.0%"),
    ("Conversion Rate %", f"=SUM({RD}!{OC}2:{OC}{LAST_ROW})/{N}", "0.0%"),
    ("Return Rate %", f'=COUNTIF({RD}!{RET}2:{RET}{LAST_ROW},"Yes")/SUM({RD}!{OC}2:{OC}{LAST_ROW})', "0.0%"),
]
positions = [("A4","B4"),("A6","B6"),("A8","B8"),("A10","B10"),
             ("D4","E4"),("D6","E6"),("D8","E8"),("D10","E10")]
for (label, formula, numfmt), (lbl_cell, val_cell) in zip(kpis, positions):
    kd[lbl_cell] = label
    kd[lbl_cell].font = LABEL_FONT
    vcell = kd[val_cell]
    vcell.value = formula
    vcell.number_format = numfmt
    vcell.font = Font(name=FONT, bold=True, size=13, color="1F4E78")
    kd[lbl_cell].fill = KPI_FILL
    vcell.fill = KPI_FILL
    kd[lbl_cell].border = BORDER
    vcell.border = BORDER

for c, w in zip("ABCDE", [22,16,4,22,16]):
    kd.column_dimensions[c].width = w
for r in range(4, 12, 2):
    kd.row_dimensions[r].height = 22

wb.save(OUT)
print(f"Workbook written: {OUT} — total build time {time.time()-t0:.1f}s")
