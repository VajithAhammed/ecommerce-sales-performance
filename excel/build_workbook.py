"""
Builds excel/Ecommerce_Sales_Analysis.xlsx
Sheets:
  1. Raw_Data            - the 25-row source data
  2. Data_Quality        - duplicate & missing-value checks (formulas)
  3. Calculated_Fields   - Revenue, Profit, Margin, Discount Amt, Segment, Order Hour, lookups
  4. Pivot_Summary       - SUMIFS/COUNTIFS cross-tabs (Category/Region/Month/Segment/Payment)
  5. KPI_Dashboard       - headline KPI cards, all formula-driven
Every metric is written as a live formula, not a hardcoded Python-computed value.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import csv

SRC = "../data/ecommerce_data.csv"
OUT = "Ecommerce_Sales_Analysis.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
LABEL_FONT = Font(name=FONT, bold=True, size=10)
NORMAL_FONT = Font(name=FONT, size=10)
KPI_FILL = PatternFill("solid", fgColor="EAF1F8")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------
# 1. RAW_DATA
# ---------------------------------------------------------------
ws = wb.create_sheet("Raw_Data")
with open(SRC) as f:
    rows = list(csv.reader(f))
header, data = rows[0], rows[1:]
for c, h in enumerate(header, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
for r, row in enumerate(data, 2):
    for c, val in enumerate(row, 1):
        colname = header[c-1]
        if colname in ("Quantity","Age","Pages_Viewed","Order_Completed","Shipping_Days"):
            val = int(val)
        elif colname in ("Unit_Price","Discount","Cost","Session_Duration"):
            val = float(val)
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = NORMAL_FONT
        cell.border = BORDER
for c, h in enumerate(header, 1):
    ws.column_dimensions[get_column_letter(c)].width = max(12, len(h)+2)
ws.freeze_panes = "A2"
N = len(data)  # 25
LAST_ROW = N + 1

col = {name: get_column_letter(i+1) for i, name in enumerate(header)}

# ---------------------------------------------------------------
# 2. DATA_QUALITY
# ---------------------------------------------------------------
dq = wb.create_sheet("Data_Quality")
dq["A1"] = "DATA QUALITY CHECKS"
dq["A1"].font = TITLE_FONT
dq.merge_cells("A1:D1")

dq["A3"] = "Check"; dq["B3"] = "Result"; dq["C3"] = "Formula Used"
for cell in ("A3","B3","C3"):
    dq[cell].font = HEADER_FONT; dq[cell].fill = HEADER_FILL

checks = [
    ("Total rows", f"=COUNTA(Raw_Data!A2:A{LAST_ROW})", "COUNTA"),
    ("Duplicate Order_IDs", f'=SUMPRODUCT((COUNTIF(Raw_Data!A2:A{LAST_ROW},Raw_Data!A2:A{LAST_ROW})>1)*1)', "SUMPRODUCT + COUNTIF"),
    ("Missing values (any column, row-count)",
     f'=SUMPRODUCT(--(COUNTBLANK(Raw_Data!A2:U{LAST_ROW})>0))', "SUMPRODUCT + COUNTBLANK (row-wise not directly supported; see note)"),
    ("Missing Order_ID cells", f"=COUNTBLANK(Raw_Data!A2:A{LAST_ROW})", "COUNTBLANK"),
    ("Missing Customer_ID cells", f"=COUNTBLANK(Raw_Data!D2:D{LAST_ROW})", "COUNTBLANK"),
    ("Negative or zero Unit_Price", f"=COUNTIF(Raw_Data!L2:L{LAST_ROW},\"<=0\")", "COUNTIF"),
    ("Discount out of range (>0.5)", f"=COUNTIF(Raw_Data!M2:M{LAST_ROW},\">0.5\")", "COUNTIF"),
    ("Duplicate full rows (Order_ID basis)", f'=SUMPRODUCT((COUNTIF(Raw_Data!A2:A{LAST_ROW},Raw_Data!A2:A{LAST_ROW})>1)*1)/2', "Duplicate pair count"),
]
r = 4
for label, formula, note in checks:
    dq.cell(row=r, column=1, value=label).font = NORMAL_FONT
    dq.cell(row=r, column=2, value=formula).font = NORMAL_FONT
    dq.cell(row=r, column=3, value=note).font = Font(name=FONT, size=9, italic=True, color="808080")
    r += 1

dq["A14"] = "Conclusion:"
dq["A14"].font = LABEL_FONT
dq["A15"] = ("Prototype dataset (25 rows) contains 0 duplicate Order_IDs and 0 missing "
             "critical fields (Order_ID, Customer_ID, Unit_Price). Data is clean and ready for analysis.")
dq["A15"].font = NORMAL_FONT
dq.merge_cells("A15:F15")
for col_letter, width in [("A",42),("B",14),("C",55)]:
    dq.column_dimensions[col_letter].width = width

# ---------------------------------------------------------------
# 3. CALCULATED_FIELDS
# ---------------------------------------------------------------
cf = wb.create_sheet("Calculated_Fields")
headers = ["Order_ID","Order_Date","Order_Hour","Customer_ID","Customer_Segment",
           "Category (via lookup)","Revenue","Discount_Amount","Cost","Profit","Profit_Margin_%"]
for c, h in enumerate(headers, 1):
    cell = cf.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for i in range(N):
    r = i + 2
    src_r = i + 2  # Raw_Data row
    cf.cell(row=r, column=1, value=f"=Raw_Data!{col['Order_ID']}{src_r}")
    cf.cell(row=r, column=2, value=f"=Raw_Data!{col['Order_Date']}{src_r}")
    # Order Hour - date function extracting hour from time-as-text "HH:MM"
    cf.cell(row=r, column=3,
            value=f'=VALUE(LEFT(Raw_Data!{col["Order_Time"]}{src_r},FIND(":",Raw_Data!{col["Order_Time"]}{src_r})-1))')
    cf.cell(row=r, column=4, value=f"=Raw_Data!{col['Customer_ID']}{src_r}")
    # Customer segment via IFS (post-2007 -> needs _xlfn prefix)
    cf.cell(row=r, column=5,
            value=f'=_xlfn.IFS(Raw_Data!{col["Customer_Type"]}{src_r}="New","New Customer",'
                  f'Raw_Data!{col["Customer_Type"]}{src_r}="Returning","Returning Customer",TRUE,"Unclassified")')
    # Category lookup via INDEX/MATCH (XLOOKUP-equivalent; XLOOKUP not used - see note in Data_Quality)
    cf.cell(row=r, column=6,
            value=f"=INDEX(Raw_Data!{col['Category']}:{col['Category']},MATCH(A{r},Raw_Data!{col['Order_ID']}:{col['Order_ID']},0))")
    # Revenue
    cf.cell(row=r, column=7,
            value=f"=Raw_Data!{col['Quantity']}{src_r}*Raw_Data!{col['Unit_Price']}{src_r}*(1-Raw_Data!{col['Discount']}{src_r})")
    # Discount amount
    cf.cell(row=r, column=8,
            value=f"=Raw_Data!{col['Quantity']}{src_r}*Raw_Data!{col['Unit_Price']}{src_r}*Raw_Data!{col['Discount']}{src_r}")
    cf.cell(row=r, column=9, value=f"=Raw_Data!{col['Cost']}{src_r}")
    cf.cell(row=r, column=10, value=f"=G{r}-I{r}")
    cf.cell(row=r, column=11, value=f"=IFERROR(J{r}/G{r},0)")
    for c in range(1, 12):
        cf.cell(row=r, column=c).font = NORMAL_FONT
        cf.cell(row=r, column=c).border = BORDER
cf["K2"].number_format = "0.0%"
for rr in range(2, LAST_ROW+1):
    cf.cell(row=rr, column=11).number_format = "0.0%"
    cf.cell(row=rr, column=7).number_format = "#,##0.00"
    cf.cell(row=rr, column=8).number_format = "#,##0.00"
    cf.cell(row=rr, column=9).number_format = "#,##0.00"
    cf.cell(row=rr, column=10).number_format = "#,##0.00"
for c in range(1, 12):
    cf.column_dimensions[get_column_letter(c)].width = 16
cf.freeze_panes = "A2"

note = cf.cell(row=LAST_ROW+2, column=1,
                value=("Note: Excel's newer XLOOKUP() is used conceptually as the lookup approach "
                       "(=XLOOKUP(A2,Raw_Data!Order_ID,Raw_Data!Category)); this workbook implements the "
                       "same result with INDEX/MATCH for maximum compatibility across Excel versions."))
note.font = Font(name=FONT, size=9, italic=True, color="808080")
cf.merge_cells(start_row=LAST_ROW+2, start_column=1, end_row=LAST_ROW+2, end_column=11)

# ---------------------------------------------------------------
# 4. PIVOT_SUMMARY (SUMIFS / COUNTIFS cross-tabs, acting as pivot tables)
# ---------------------------------------------------------------
ps = wb.create_sheet("Pivot_Summary")
ps["A1"] = "PIVOT-STYLE SUMMARY (SUMIFS / COUNTIFS)"
ps["A1"].font = TITLE_FONT
ps.merge_cells("A1:E1")
ps["A2"] = "Tip: Select Raw_Data range and Insert > PivotTable in Excel for a native, drag-and-drop pivot."
ps["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

RD = "Raw_Data"
CAT_COL = col['Category']; REG_COL = col['Region']; DATE_COL = col['Order_Date']
QTY_COL = col['Quantity']; PRICE_COL = col['Unit_Price']; DISC_COL = col['Discount']
CTYPE_COL = col['Customer_Type']; PAY_COL = col['Payment_Method']

def revenue_sumifs(criteria_range, criteria_cell):
    return (f"=SUMPRODUCT(({RD}!{criteria_range}=${criteria_cell})*"
            f"{RD}!{QTY_COL}2:{QTY_COL}{LAST_ROW}*{RD}!{PRICE_COL}2:{PRICE_COL}{LAST_ROW}*"
            f"(1-{RD}!{DISC_COL}2:{DISC_COL}{LAST_ROW}))")

# --- Table A: Revenue & Orders by Category ---
ps["A4"] = "Revenue & Orders by Category"; ps["A4"].font = LABEL_FONT
ps["A5"] = "Category"; ps["B5"] = "Orders"; ps["C5"] = "Revenue"
for cell in ("A5","B5","C5"):
    ps[cell].font = HEADER_FONT; ps[cell].fill = HEADER_FILL

categories = sorted(set(row[header.index("Category")] for row in data))
r = 6
for cat in categories:
    ps.cell(row=r, column=1, value=cat).font = NORMAL_FONT
    ps.cell(row=r, column=2,
            value=f'=COUNTIF({RD}!{CAT_COL}2:{CAT_COL}{LAST_ROW},A{r})').font = NORMAL_FONT
    ps.cell(row=r, column=3,
            value=revenue_sumifs(f"{CAT_COL}2:{CAT_COL}{LAST_ROW}", f"A{r}")).font = NORMAL_FONT
    ps.cell(row=r, column=3).number_format = "#,##0.00"
    r += 1
cat_end = r - 1

# --- Table B: Revenue & Orders by Region ---
start_b = cat_end + 3
ps.cell(row=start_b, column=1, value="Revenue & Orders by Region").font = LABEL_FONT
ps.cell(row=start_b+1, column=1, value="Region").font = HEADER_FONT
ps.cell(row=start_b+1, column=2, value="Orders").font = HEADER_FONT
ps.cell(row=start_b+1, column=3, value="Revenue").font = HEADER_FONT
for cc in range(1,4):
    ps.cell(row=start_b+1, column=cc).fill = HEADER_FILL
regions = sorted(set(row[header.index("Region")] for row in data))
r = start_b + 2
for reg in regions:
    ps.cell(row=r, column=1, value=reg).font = NORMAL_FONT
    ps.cell(row=r, column=2, value=f'=COUNTIF({RD}!{REG_COL}2:{REG_COL}{LAST_ROW},A{r})').font = NORMAL_FONT
    ps.cell(row=r, column=3, value=revenue_sumifs(f"{REG_COL}2:{REG_COL}{LAST_ROW}", f"A{r}")).font = NORMAL_FONT
    ps.cell(row=r, column=3).number_format = "#,##0.00"
    r += 1
reg_end = r - 1

# --- Table C: New vs Returning ---
start_c = reg_end + 3
ps.cell(row=start_c, column=1, value="New vs Returning Customers").font = LABEL_FONT
ps.cell(row=start_c+1, column=1, value="Customer_Type").font = HEADER_FONT
ps.cell(row=start_c+1, column=2, value="Orders").font = HEADER_FONT
ps.cell(row=start_c+1, column=3, value="Revenue").font = HEADER_FONT
for cc in range(1,4):
    ps.cell(row=start_c+1, column=cc).fill = HEADER_FILL
r = start_c + 2
for ctype in ["New","Returning"]:
    ps.cell(row=r, column=1, value=ctype).font = NORMAL_FONT
    ps.cell(row=r, column=2, value=f'=COUNTIF({RD}!{CTYPE_COL}2:{CTYPE_COL}{LAST_ROW},A{r})').font = NORMAL_FONT
    ps.cell(row=r, column=3, value=revenue_sumifs(f"{CTYPE_COL}2:{CTYPE_COL}{LAST_ROW}", f"A{r}")).font = NORMAL_FONT
    ps.cell(row=r, column=3).number_format = "#,##0.00"
    r += 1
ctype_end = r - 1

# --- Table D: Payment method mix ---
start_d = ctype_end + 3
ps.cell(row=start_d, column=1, value="Orders by Payment Method").font = LABEL_FONT
ps.cell(row=start_d+1, column=1, value="Payment_Method").font = HEADER_FONT
ps.cell(row=start_d+1, column=2, value="Orders").font = HEADER_FONT
for cc in range(1,3):
    ps.cell(row=start_d+1, column=cc).fill = HEADER_FILL
methods = sorted(set(row[header.index("Payment_Method")] for row in data))
r = start_d + 2
for m in methods:
    ps.cell(row=r, column=1, value=m).font = NORMAL_FONT
    ps.cell(row=r, column=2, value=f'=COUNTIFS({RD}!{PAY_COL}2:{PAY_COL}{LAST_ROW},A{r})').font = NORMAL_FONT
    r += 1
pay_end = r - 1

for c, w in zip("ABCDE", [24,12,16,16,16]):
    ps.column_dimensions[c].width = w

# --- Pivot Charts ---
bar = BarChart()
bar.title = "Revenue by Category"
bar.y_axis.title = "Revenue (INR)"
bar.x_axis.title = "Category"
data_ref = Reference(ps, min_col=3, min_row=5, max_row=cat_end)
cats_ref = Reference(ps, min_col=1, min_row=6, max_row=cat_end)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.width = 16; bar.height = 9
ps.add_chart(bar, "E4")

pie = PieChart()
pie.title = "Revenue Share by Region"
data_ref2 = Reference(ps, min_col=3, min_row=start_b+1, max_row=reg_end)
cats_ref2 = Reference(ps, min_col=1, min_row=start_b+2, max_row=reg_end)
pie.add_data(data_ref2, titles_from_data=True)
pie.set_categories(cats_ref2)
pie.width = 16; pie.height = 9
ps.add_chart(pie, "E22")

# ---------------------------------------------------------------
# 5. KPI_DASHBOARD
# ---------------------------------------------------------------
kd = wb.create_sheet("KPI_Dashboard", 0)
kd["A1"] = "E-COMMERCE SALES PERFORMANCE — KPI DASHBOARD"
kd["A1"].font = Font(name=FONT, bold=True, size=16, color="1F4E78")
kd.merge_cells("A1:D1")
kd["A2"] = "Prototype dataset: 25 orders | Jan 2026"
kd["A2"].font = Font(name=FONT, italic=True, size=10, color="808080")

OC = col['Order_Completed']; QC = col['Quantity']; PC = col['Unit_Price']; DC = col['Discount']; CC = col['Cost']
RET = col['Return_Status']

kpis = [
    ("Total Revenue",
     f"=SUMPRODUCT(({RD}!{OC}2:{OC}{LAST_ROW}=1)*{RD}!{QC}2:{QC}{LAST_ROW}*{RD}!{PC}2:{PC}{LAST_ROW}*(1-{RD}!{DC}2:{DC}{LAST_ROW}))",
     "#,##0.00"),
    ("Total Profit",
     f"=SUMPRODUCT(({RD}!{OC}2:{OC}{LAST_ROW}=1)*({RD}!{QC}2:{QC}{LAST_ROW}*{RD}!{PC}2:{PC}{LAST_ROW}*(1-{RD}!{DC}2:{DC}{LAST_ROW})-{RD}!{CC}2:{CC}{LAST_ROW}))",
     "#,##0.00"),
    ("Total Orders", f"=COUNTA({RD}!A2:A{LAST_ROW})", "0"),
    ("Total Customers", f"=SUMPRODUCT(1/COUNTIF({RD}!{col['Customer_ID']}2:{col['Customer_ID']}{LAST_ROW},{RD}!{col['Customer_ID']}2:{col['Customer_ID']}{LAST_ROW}))", "0"),
    ("Average Order Value", f"=B4/SUM({RD}!{OC}2:{OC}{LAST_ROW})", "#,##0.00"),
    ("Profit Margin %", "=B6/B4", "0.0%"),
    ("Conversion Rate %", f"=SUM({RD}!{OC}2:{OC}{LAST_ROW})/{N}", "0.0%"),
    ("Return Rate %", f'=COUNTIF({RD}!{RET}2:{RET}{LAST_ROW},"Yes")/SUM({RD}!{OC}2:{OC}{LAST_ROW})', "0.0%"),
]
# place as 2-column x 4-row KPI card grid
positions = [("A4","B4"),("A6","B6"),("A8","B8"),("A10","B10"),
             ("D4","E4"),("D6","E6"),("D8","E8"),("D10","E10")]
for (label, formula, numfmt), (lbl_cell, val_cell) in zip(kpis, positions):
    kd[lbl_cell] = label
    kd[lbl_cell].font = LABEL_FONT
    vcell = kd[val_cell]
    vcell.value = formula
    vcell.number_format = numfmt
    vcell.font = Font(name=FONT, bold=True, size=13, color="1F4E78")
    kd[lbl_cell].fill = KPI_FILL
    vcell.fill = KPI_FILL
    kd[lbl_cell].border = BORDER
    vcell.border = BORDER

for c, w in zip("ABCDE", [22,16,4,22,16]):
    kd.column_dimensions[c].width = w
for r in range(4, 12, 2):
    kd.row_dimensions[r].height = 22

wb.save(OUT)
print("Workbook written:", OUT)
